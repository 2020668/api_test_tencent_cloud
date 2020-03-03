# -*- coding: utf-8 -*-
"""

=================================
Author: keen
Created on: 2019/6/4

E-mail:keen2020@outlook.com

=================================


"""

import openpyxl


class Case(object):
    # 用这个类来存储用例 attrs为一个zip对象 excel的sheet中有几行 就返回几个zip对象
    def __init__(self, attrs):
        """
        初始化用例
        :param attrs: zip类型--> [（key1,value1),（key2,value2)....]
        """
        # item遍历所有zip对象 并获取每个zip对象内部的所有元素 返回tuple example: ('title', '正常登录')
        for item in attrs:
            # 将key-value组装起来
            setattr(self, item[0], item[1])


class ReadExcel(object):
    """
    读取Excel数据
    """

    def __init__(self, file_name, sheet_name):
        """
        初始化读取对象
        :param file_name: 文件名，测试用例文件--> str
        :param sheet_name: 表单名--> str
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        self.wb = openpyxl.load_workbook(self.file_name)  # 打开工作簿，传入的指定文件名
        self.sheet = self.wb[self.sheet_name]    # 选取表单，传入的指定表单

    def close(self):
        self.wb.close()

    def read_line_data(self):
        """
        执行读取数据
        :return: list
        """
        # 打开工作簿
        self.open()
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)

        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            # 对title是否为空进行过滤，容错机制
            if title.value:
                titles.append(title.value)

        # 定义一个空列表用来存储所有的用例
        cases = []
        # 从第2行开始，就是测试用例数据了
        for case in rows_data[1:]:
            # data用例临时存放每行的用例数据
            data = []
            for cell in case:
                data.append(cell.value)
            # 将title与测试用例数据组合，形成每行测试用例,一行行读取数据，无需加list
            case_data = zip(titles, data)
            # 创建一个Case类的对象，用来保存用例数据，
            case_obj = Case(case_data)
            # 将该条数据放入cases中
            cases.append(case_obj)
        # 关闭工作簿,并返回读取的数据-->list
        self.close()
        return cases

    def read_column_data(self, read_column):
        """
        执行读取数据
        :param read_column:指定读取的列[1,2,3,4...]
        :return: 返回一个list，每个元素为一个用例,dict
        """
        # 打开工作簿
        self.open()
        if len(read_column) == 0:
            return self.read_line_data()

        # 获取该sheet下最大的行数 就是有多少行数据
        max_r = self.sheet.max_row
        cases = []  # 存储所有用例
        titles = []  # 存放标题
        case_data = []  # 存储该行的数据
        # 遍历该sheet下所有的行
        for row in range(1, max_r+1):
            if row > 1:    # 排除表头，获取数据
                # case_data = []      # 存储该行的数据
                for column in read_column:
                    case_rc = self.sheet.cell(row, column).value   # 获取指定单元格的数据
                    case_data.append(case_rc)
                case = zip(titles, case_data)  # 将表头和数据进行打包，--> 第1次遍历已经添加了表头,一行行读取数据，故无需加list
                case_obj = Case(case)
                cases.append(case_obj)
            else:   # 获取表头title信息
                # read_column是一个列表 [1, 3]意思是只读取第1  第3列的数据
                for column in read_column:
                    # 传入指定的列 与从1开始的行组合 获取具体的value
                    title = self.sheet.cell(row, column).value
                    # 如果读取到的值不为空 则将该值添加到列表titles中
                    if title:
                        titles.append(title)
        # 关闭工作簿
        self.close()
        return cases

    def write_data(self, row, column, value):
        # 打开工作簿
        self.open()
        # 指定位置写入数据
        self.sheet.cell(row=row, column=column, value=value)
        # 保存数据
        self.wb.save(self.file_name)
        # 关闭工作簿
        self.close()


if __name__ == '__main__':
    import os
    from common.constant import DATA_DIR
    from common.config import conf
    # 从配置文件获取数据
    file_name = conf.get('excel', 'file_name')
    read_column = conf.get('excel', 'read_column')
    read_column = eval(read_column)  # 将str转换成list
    wb = ReadExcel(os.path.join(DATA_DIR, file_name), "login")
    # 返回list 每个元素（对象）对应excel中一行数据 对象对应的属性值 对应excel中每行title下的值
    # cases = wb.read_column_data(read_column)
    cases = wb.read_line_data()
    for case in cases:
        print(case.title, case.url, case.request_data)

    # 在13行3列写入
    wb.write_data(13, 3, "哈哈")

# import openpyxl
#
#
# class Case(object):
#     def __init__(self, attrs):
#         for item in attrs:
#             setattr(self, item[0], item[1])
#
#
# class ReadExcel(object):
#     def __init__(self, file_name, sheet_name):
#         self.file_name = file_name
#         self.sheet_name = sheet_name
#
#     def open(self):
#         self.wb = openpyxl.load_workbook(self.file_name)
#         self.sheet = self.wb[self.sheet_name]
#
#     def close(self):
#         self.wb.close()
#
#     def read_data(self):
#         self.open()
#         rows_data = list(self.sheet.rows)
#         cases = []
#
#         titles = []
#         for title in rows_data[0]:
#             if title.value:
#                 titles.append(title.value)
#
#         for case in rows_data[1:]:
#             data = []
#             for cell in case:
#                 data.append(cell.value)
#             case_data = zip(titles, data)
#             case_obj = Case(case_data)
#             cases.append(case_obj)
#
#         self.close()
#         return cases
#
#     def read_column_data(self, read_column):
#         if len(read_column) == 0:
#             return self.read_data()
#         self.open()
#         cases = []
#         titles = []
#         max_r = self.sheet.max_row
#         for row in range(1, max_r+1):
#             if row != 1:
#                 data = []
#                 for column in read_column:
#                     data.append(self.sheet.cell(row, column).value)
#                 case_data = zip(titles, data)
#                 case_obj = Case(case_data)
#                 cases.append(case_obj)
#             else:
#                 for column in read_column:
#                     title = self.sheet.cell(row, column).value
#                     if title is not None:
#                         titles.append(title)
#
#         self.close()
#         return cases
#
#     def write_data(self, row, column, value):
#         self.open()
#         self.sheet.cell(row=row, column=column, value=value)
#         self.wb.save(self.file_name)
#         self.close()

